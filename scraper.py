import requests
from bs4 import BeautifulSoup
import csv
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


BASE_URL = "http://books.toscrape.com/"


def get_page(url):
    headers = {
        "User-Agent": "Mozilla/5.0"
    }
    response = requests.get(url, headers=headers, timeout=10)
    response.raise_for_status()
    return response.text


def parse_books(html):
    soup = BeautifulSoup(html, "html.parser")
    books = []

    for book in soup.select("article.product_pod"):
        title = book.h3.a["title"]
        price = book.select_one(".price_color").text
        if price[0] == "Â":
            price = price[1:]
        rating = book.p["class"][1]

        books.append({
            "title": title,
            "price": price,
            "rating": rating
        })

    return books


def save_to_csv(data, filename="output/books.csv"):
    with open(filename, "w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=["title", "price", "rating"])
        writer.writeheader()
        writer.writerows(data)

def to_excel(books, filename):
    df = pd.DataFrame(books)
    df.to_excel(filename, index=False)
    wb = load_workbook(filename)
    ws = wb.active
    for column in ws.columns:
        max_len = max(len(str(cell.value)) for cell in column)
        adjusted_len = min(max_len, 40)
        new_len = max(max_len,10)
        ws.column_dimensions[column[0].column_letter].width = adjusted_len
        ws.column_dimensions[column[1].column_letter].width = new_len
        ws.column_dimensions[column[2].column_letter].width = new_len

    for cell in ws['A']:
        cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

    for cell in ws['B']:
        cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

    for cell in ws['C']:
        cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
    for cell in ws[1]:
        cell.font = Font(bold=True)
    wb.save(filename)


def main():
    html = get_page(BASE_URL)
    books = parse_books(html)
    save_to_csv(books)
    to_excel(books, "output/books.xlsx")
    print("Scraping completed successfully!")


if __name__ == "__main__":
    main()
